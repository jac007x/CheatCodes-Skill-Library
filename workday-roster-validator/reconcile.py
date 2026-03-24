#!/usr/bin/env python3
"""
workday-roster-validator — reconcile.py  v1.1.0
=================================================
Validates whether Workday Worker Change History proposed values
have propagated into a downstream ARS/PowerBI roster export.

Usage
-----
  python3 reconcile.py \
    --changes  /path/to/Worker_Change_History_Report.xlsx \
    --roster   /path/to/ARS_Roster_MMDDYY_Power_BI.xlsx \
    [--snapshot-date 2026-03-24] \
    [--out      /path/to/report.xlsx] \
    [--no-location] \
    [--fields  "Manager(s) - Proposed,Cost Center - Proposed"]

Result Classifications
----------------------
  ✅ Match    — proposed value (normalized) equals roster value
  ⚠️ Mismatch — proposed value does not match roster (change not reflected)
  ⏩ No Change — proposed == current (field wasn't actually changed)
  🕐 Pending  — effective date is after the roster snapshot date (expected)
  ❌ Not Found — WIN not in roster

Assumptions
-----------
  Change report: .xlsx, first sheet
    Row 0: report title ("Work History Summary") — skipped
    Row 1: column headers
    Row 2+: data rows
    Join key: "Associate ID"

  Roster: .xlsx, first sheet
    Row 0: column headers (single header row)
    Row 1+: data rows
    Join key: "WIN Number"
"""

import re
import sys
import argparse
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path
from typing import Optional, Set

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── US state codes for location parsing ───────────────────────────────────────
US_STATES = (
    'AL|AK|AZ|AR|CA|CO|CT|DE|FL|GA|HI|ID|IL|IN|IA|KS|KY|LA|ME|MD|MA|MI|MN|'
    'MS|MO|MT|NE|NV|NH|NJ|NM|NY|NC|ND|OH|OK|OR|PA|RI|SC|SD|TN|TX|UT|VT|VA|'
    'WA|WV|WI|WY'
)

# ── Styles ─────────────────────────────────────────────────────────────────────
HDR_FILL     = PatternFill("solid", fgColor="1F2D5B")
HDR_FONT     = Font(bold=True, color="FFFFFF", size=10)
OK_FILL      = PatternFill("solid", fgColor="C6EFCE")   # green
WARN_FILL    = PatternFill("solid", fgColor="FFEB9C")   # yellow
ERR_FILL     = PatternFill("solid", fgColor="FFC7CE")   # red
PENDING_FILL = PatternFill("solid", fgColor="DDEEFF")   # blue — pending/no-change
SKIP_FILL    = PatternFill("solid", fgColor="F2F2F2")   # grey — no change
BLUE_FILL    = PatternFill("solid", fgColor="DDEEFF")
THIN         = Side(style='thin', color="CCCCCC")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SEVERITY = {
    'Manager(s) - Proposed → Manager Name':       '🔴 High',
    'Cost Center - Proposed → Cost Center #':      '🔴 High',
    'Job Code - Proposed → Job Code':              '🟡 Medium',
    'Job Profile - Proposed → Job Description':    '🟡 Medium',
    'Position - Proposed → Position Description':  '🟡 Medium',
    'Location - Proposed → Work City/State':       '🟢 Low',
}
SEV_ORDER = {'🔴 High': 0, '🟡 Medium': 1, '🟢 Low': 2}

# Match icons → fill mapping
MATCH_FILLS = {
    '✅': OK_FILL,
    '⚠️': WARN_FILL,
    '⏩': SKIP_FILL,
    '🕐': PENDING_FILL,
    '❌': ERR_FILL,
}


# ── Normalization helpers ──────────────────────────────────────────────────────

def norm(v):
    """Lowercase, collapse whitespace."""
    if v is None:
        return ''
    return re.sub(r'\s+', ' ', str(v).strip().lower())


def extract_job_code(jc):
    """Strip country prefix: 'US-100022465' → '100022465'."""
    return re.sub(r'^[A-Z]{2,3}[-_]', '', str(jc).strip()) if jc else ''


def extract_manager_name(mgr):
    """
    'ARPAN BAJORIA (BAJORIA)'  → 'ARPAN BAJORIA'
    'Ian Hanson - SC Product'  → 'Ian Hanson'
    """
    if not mgr:
        return ''
    s = re.sub(r'\s*\([^)]+\)\s*$', '', str(mgr).strip())
    return re.split(r'\s+-\s+', s)[0].strip()


def extract_cc_code(cc):
    """'US11919 11919 PD SUPPLY CHAIN FULFILLMENT' → 'US11919'."""
    return str(cc).strip().split()[0].upper() if cc else ''


def extract_position_desc(pos):
    """
    'P_0010902585 Staff, Product Manager' → 'Staff, Product Manager'
    '(USA) Staff, Product Manager'        → 'Staff, Product Manager'
    Strip P_code prefix then (COUNTRY) prefix from both sides.
    """
    if not pos:
        return ''
    s = str(pos).strip()
    m = re.match(r'^P_\d+\s+(.+)$', s)
    if m:
        s = m.group(1)
    s = re.sub(r'^\([A-Z]{2,3}\)\s+', '', s)
    return s.strip()


def extract_wd_location(loc):
    """
    '(USA) Change Building AR Bentonville Home Office' → ('BENTONVILLE', 'AR')
    Algorithm: strip 'Home Office', find last US state code + following city text.
    """
    if not loc:
        return '', ''
    s = str(loc).strip().upper()
    s = re.sub(r'\s+HOME OFFICE\s*$', '', s).strip()
    matches = list(re.finditer(r'\b(' + US_STATES + r')\s+([A-Z][A-Z0-9\s\-\'\.]+)', s))
    if matches:
        last = matches[-1]
        return last.group(2).strip(), last.group(1)
    return '', ''


def parse_date(v) -> Optional[date]:
    """Parse a cell value to a date. Returns None on failure."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.date() if isinstance(v, datetime) else v
    try:
        return datetime.strptime(str(v)[:10], '%Y-%m-%d').date()
    except ValueError:
        return None


# ── Field mapping ──────────────────────────────────────────────────────────────
# (proposed_col, current_col, roster_col, proposed_norm_fn, roster_norm_fn)
BASE_FIELD_MAP = [
    ('Job Code - Proposed',    'Job Code - Current',    'Job Code',             extract_job_code,      extract_job_code),
    ('Job Profile - Proposed', 'Job Profile - Current', 'Job Description',      str.strip,             str.strip),
    ('Manager(s) - Proposed',  'Manager(s) - Current',  'Manager Name',         extract_manager_name,  extract_manager_name),
    ('Cost Center - Proposed', 'Cost Center - Current', 'Cost Center #',        extract_cc_code,       norm),
    ('Location - Proposed',    'Location - Current',    'Work City/State',      None,                  None),   # special
    ('Position - Proposed',    'Position - Current',    'Position Description', extract_position_desc, extract_position_desc),
]


# ── Loaders ────────────────────────────────────────────────────────────────────

def load_roster(path: str) -> dict:
    """Returns {win: {col: value}} for the full roster."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    win_idx = headers.index('WIN Number')
    roster = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= win_idx:
            continue
        win = str(row[win_idx]).strip() if row[win_idx] else None
        if win and win != 'None':
            padded = list(row) + [None] * (len(headers) - len(row))
            roster[win] = dict(zip(headers, padded))
    wb.close()
    return roster


def load_changes(path: str) -> dict:
    """
    Returns {win: [row_dict, ...]} for the change report.
    Skips row 0 (report title), uses row 1 as headers.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    headers = [str(v).strip() if v else '' for v in all_rows[1]]
    win_changes = defaultdict(list)
    for row in all_rows[2:]:
        if not any(v for v in row):
            continue
        padded = list(row) + [None] * (len(headers) - len(row))
        d = dict(zip(headers, padded))
        win = str(d.get('Associate ID', '') or '').strip()
        if win:
            win_changes[win].append(d)
    wb.close()
    return dict(win_changes)


def resolve_snapshot_date(roster: dict, roster_path: str) -> Optional[date]:
    """
    Try to resolve snapshot date. Priority:
    1. Filename pattern (most reliable for ARS exports, e.g. 3.24.26)
    2. Snapshot Date column in roster data (may lag behind filename)
    """
    # 1. Try from filename: match patterns like 3.24.26 or 2026-03-24
    fname = Path(roster_path).stem
    m = re.search(r'(\d{1,2})[.\-_](\d{1,2})[.\-_](\d{2,4})', fname)
    if m:
        mo, day, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if yr < 100:
            yr += 2000
        try:
            return date(yr, mo, day)
        except ValueError:
            pass
    # 2. Try from roster Snapshot Date column
    sample = next(iter(roster.values()), {})
    d = parse_date(sample.get('Snapshot Date'))
    if d:
        return d
    return None


# ── Core reconciliation ────────────────────────────────────────────────────────

def reconcile(
    roster: dict,
    win_changes: dict,
    snapshot_date: Optional[date] = None,
    include_location: bool = True,
    fields_filter: Optional[Set] = None,
) -> list:
    """
    For each WIN in win_changes, compare all proposed field values against
    the roster. Returns a list of result dicts.

    Result classifications per field:
      ✅  Match    — normalized proposed == normalized roster value
      ⚠️  Mismatch — proposed does not match roster
      ⏩  No Change — proposed == current (field not actually changed)
      🕐  Pending  — effective date after snapshot_date (not yet expected)
    """
    # Build active field map (filter by fields_filter + include_location)
    field_map = []
    for entry in BASE_FIELD_MAP:
        prop_col = entry[0]
        if prop_col == 'Location - Proposed' and not include_location:
            continue
        if fields_filter and prop_col not in fields_filter:
            continue
        field_map.append(entry)

    results = []

    for win, chg_rows in sorted(win_changes.items()):
        name = chg_rows[0].get('Legal Name in General Display Format', '')
        bp_types = ', '.join(sorted({c.get('Business Process Type', '') for c in chg_rows}))
        eff_dates_raw = [parse_date(c.get('Effective Date')) for c in chg_rows]
        eff_dates_clean = [d for d in eff_dates_raw if d]
        eff_dates_str = ', '.join(sorted({str(d) for d in eff_dates_clean}))
        is_hire = any('hire' in str(c.get('Business Process Type', '')).lower() for c in chg_rows)
        latest_eff = max(eff_dates_clean) if eff_dates_clean else None

        if win not in roster:
            label = '❌ NEW HIRE (expected)' if is_hire else '❌ NOT IN ROSTER'
            results.append({
                'WIN': win, 'Name': name, 'Status': label,
                'BP Types': bp_types, 'Effective Dates': eff_dates_str, 'fields': [],
            })
            continue

        roster_rec = roster[win]
        field_results = []
        all_match = True
        any_pending = False

        for prop_col, curr_col, roster_col, prop_fn, roster_fn in field_map:

            # Collect unique proposed + current values across all rows
            proposed_set = {
                str(c[prop_col]).strip()
                for c in chg_rows
                if c.get(prop_col) is not None and str(c.get(prop_col, '')).strip() not in ('', 'None')
            }
            current_set = {
                str(c[curr_col]).strip()
                for c in chg_rows
                if c.get(curr_col) is not None and str(c.get(curr_col, '')).strip() not in ('', 'None')
            }

            if not proposed_set:
                continue  # field absent in change report — skip

            # ── ⏩ No Change: all proposed values equal all current values ──
            if proposed_set and current_set and proposed_set == current_set:
                field_results.append({
                    'field': f'{prop_col} → {roster_col}',
                    'proposed': ' | '.join(sorted(proposed_set)),
                    'roster': str(roster_rec.get(roster_col, '') or ''),
                    'match': '⏩',
                    'note': 'No change (proposed = current)',
                })
                continue

            # ── 🕐 Pending: effective date after snapshot ──
            if snapshot_date and latest_eff and latest_eff > snapshot_date:
                any_pending = True
                field_results.append({
                    'field': f'{prop_col} → {roster_col}',
                    'proposed': ' | '.join(sorted(proposed_set)),
                    'roster': str(roster_rec.get(roster_col, '') or ''),
                    'match': '🕐',
                    'note': f'Pending — effective {latest_eff} after snapshot {snapshot_date}',
                })
                continue

            # ── Location: special comparison ──
            if prop_col == 'Location - Proposed':
                r_city = str(roster_rec.get('Work City', '') or '')
                r_state = str(roster_rec.get('Work State', '') or '')
                match = any(
                    extract_wd_location(p)[0].upper() == r_city.upper() and
                    extract_wd_location(p)[1].upper() == r_state.upper()
                    for p in proposed_set
                )
                if not match:
                    all_match = False
                field_results.append({
                    'field': 'Location - Proposed → Work City/State',
                    'proposed': ' | '.join(sorted(proposed_set)),
                    'roster': f"{r_city}, {r_state}",
                    'match': '✅' if match else '⚠️',
                    'note': '',
                })
                continue

            # ── Standard field comparison ──
            roster_raw = roster_rec.get(roster_col)
            roster_norm = norm(roster_fn(roster_raw)) if roster_fn and roster_raw is not None else norm(roster_raw)
            match = any(
                norm(prop_fn(p)) == roster_norm
                for p in proposed_set
                if prop_fn(p)
            )
            if not match:
                all_match = False
            field_results.append({
                'field': f'{prop_col} → {roster_col}',
                'proposed': ' | '.join(sorted(proposed_set)),
                'roster': str(roster_raw) if roster_raw is not None else '',
                'match': '✅' if match else '⚠️',
                'note': '',
            })

        # Overall status
        real_fields = [f for f in field_results if f['match'] not in ('⏩', '🕐')]
        if any_pending and not real_fields:
            status = '🕐 PENDING'
        elif all_match:
            status = '✅ FULLY APPLIED'
        else:
            status = '⚠️ PARTIAL / MISMATCH'

        results.append({
            'WIN': win, 'Name': name, 'Status': status,
            'BP Types': bp_types, 'Effective Dates': eff_dates_str, 'fields': field_results,
        })

    return results


# ── Excel report builder ───────────────────────────────────────────────────────

def build_report(results: list, out_path: str, snapshot_date: Optional[date] = None) -> None:
    wb = Workbook()

    # ── Summary ──
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.column_dimensions['A'].width = 44
    ws_sum.column_dimensions['B'].width = 22
    ws_sum.merge_cells('A1:B1')
    title = ws_sum.cell(row=1, column=1, value="Workday → ARS Roster Reconciliation")
    title.fill = HDR_FILL
    title.font = Font(bold=True, color="FFFFFF", size=13)
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.row_dimensions[1].height = 26

    matched  = sum(1 for r in results if r['Status'] == '✅ FULLY APPLIED')
    partial  = sum(1 for r in results if '⚠️' in r['Status'])
    pending  = sum(1 for r in results if '🕐' in r['Status'])
    missing  = sum(1 for r in results if '❌' in r['Status'])
    genuine  = sum(
        1 for r in results
        if any(f['match'] == '⚠️' and 'Location' not in f['field'] for f in r.get('fields', []))
    )

    snap_label = str(snapshot_date) if snapshot_date else 'unknown'
    for i, (label, val, fill) in enumerate([
        ("Roster Snapshot Date",                snap_label,    BLUE_FILL),
        ("Generated",                           datetime.now().strftime("%Y-%m-%d %H:%M"), None),
        ("",                                    "",            None),
        ("Total WINs in Change Report",         len(results),  None),
        ("✅  Fully Applied",                   matched,       OK_FILL),
        ("⚠️  Partial Match (field mismatch)",  partial,       WARN_FILL),
        ("🕐  Pending (after snapshot date)",   pending,       PENDING_FILL),
        ("❌  Not in Roster / New Hires",        missing,       ERR_FILL),
        ("",                                    "",            None),
        ("Genuine Field Mismatches",            genuine,       WARN_FILL),
    ], 2):
        c1 = ws_sum.cell(row=i, column=1, value=label)
        c2 = ws_sum.cell(row=i, column=2, value=val)
        c1.font = Font(bold=bool(label), size=10)
        c2.font = Font(size=10)
        c2.alignment = Alignment(horizontal='center')
        if fill:
            c1.fill = fill
            c2.fill = fill

    # ── Detail ──
    ws = wb.create_sheet("Reconciliation Detail")
    col_headers = ['WIN', 'Name', 'Status', 'Business Process Types', 'Effective Dates',
                   'Field', 'Proposed Value (Workday)', 'Roster Value (ARS)', 'Match', 'Note']
    col_widths   = [14, 28, 22, 44, 18, 44, 58, 35, 8, 30]
    for col, (h, w) in enumerate(zip(col_headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    row_num = 2
    for r in results:
        fields = r.get('fields', [])
        if not fields:
            for col, val in enumerate([
                r['WIN'], r.get('Name', ''), r['Status'],
                r.get('BP Types', ''), r.get('Effective Dates', ''),
                'WIN not found in roster', '—', '—', '❌', '',
            ], 1):
                c = ws.cell(row=row_num, column=col, value=val)
                c.fill = ERR_FILL
                c.border = BORDER
                c.alignment = Alignment(wrap_text=True, vertical='top')
            row_num += 1
            continue

        start = row_num
        for i, f in enumerate(fields):
            f_fill = MATCH_FILLS.get(f['match'], WARN_FILL)
            ws.cell(row=row_num, column=6, value=f['field']).fill = f_fill
            ws.cell(row=row_num, column=7, value=f['proposed']).fill = f_fill
            ws.cell(row=row_num, column=8, value=f['roster']).fill = f_fill
            ws.cell(row=row_num, column=9, value=f['match']).fill = f_fill
            ws.cell(row=row_num, column=10, value=f.get('note', '')).fill = f_fill
            for col in range(6, 11):
                ws.cell(row=row_num, column=col).border = BORDER
                ws.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True, vertical='top')
            if i == 0:
                for col, val in enumerate([
                    r['WIN'], r.get('Name', ''), r['Status'],
                    r.get('BP Types', ''), r.get('Effective Dates', ''),
                ], 1):
                    ws.cell(row=row_num, column=col, value=val)
            row_num += 1

        if len(fields) > 1:
            for col in range(1, 6):
                ws.merge_cells(start_row=start, start_column=col,
                               end_row=row_num - 1, end_column=col)
        s_fill = (OK_FILL if '✅' in r['Status'] else
                  PENDING_FILL if '🕐' in r['Status'] else
                  ERR_FILL if '❌' in r['Status'] else WARN_FILL)
        for col in range(1, 6):
            c = ws.cell(row=start, column=col)
            c.fill = s_fill
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical='top')

    # ── Action Required ──
    ws3 = wb.create_sheet("⚠️ Action Required")
    act_headers = ['WIN', 'Name', 'Field', 'Proposed (Workday)',
                   'Roster Value (ARS)', 'BP Type', 'Severity']
    act_widths   = [14, 28, 32, 55, 30, 36, 12]
    for col, (h, w) in enumerate(zip(act_headers, act_widths), 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.fill = PatternFill("solid", fgColor="C9302C")
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
        ws3.column_dimensions[get_column_letter(col)].width = w
    ws3.freeze_panes = 'A2'

    actions = []
    for r in results:
        for f in r.get('fields', []):
            # Only genuine mismatches go to Action Required (not ⏩ No Change or 🕐 Pending)
            if f['match'] == '⚠️' and 'Location' not in f['field']:
                sev = SEVERITY.get(f['field'], '🟡 Medium')
                actions.append((r, f, sev))
        if '❌' in r['Status']:
            actions.append((r, None, '🔴 High'))
    actions.sort(key=lambda x: SEV_ORDER.get(x[2], 9))

    for act_row, (r, f, sev) in enumerate(actions, 2):
        if f:
            row_vals = [r['WIN'], r.get('Name', ''), f['field'],
                        f['proposed'], f['roster'], r.get('BP Types', '')[:50], sev]
            row_fill = WARN_FILL
        else:
            row_vals = [r['WIN'], r.get('Name', ''), 'WIN not found in roster',
                        'N/A', 'Not present', r.get('BP Types', '')[:50], sev]
            row_fill = ERR_FILL
        for col, val in enumerate(row_vals, 1):
            c = ws3.cell(row=act_row, column=col, value=val)
            c.fill = row_fill
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(out_path)


# ── CLI entry point ────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='Validate Workday change history against ARS/PowerBI roster.'
    )
    parser.add_argument('--changes',       help='Path to Worker Change History Report (.xlsx)')
    parser.add_argument('--roster',        help='Path to ARS Roster export (.xlsx)')
    parser.add_argument('--out',           help='Output path for reconciliation report (.xlsx)')
    parser.add_argument('--snapshot-date', help='Roster snapshot date YYYY-MM-DD (auto-detected if omitted)')
    parser.add_argument('--no-location',   action='store_true', help='Skip location field comparison')
    parser.add_argument('--fields',        help='Comma-separated list of Proposed columns to validate (default: all)')
    args = parser.parse_args()

    changes_path = args.changes or input('Change report path: ').strip()
    roster_path  = args.roster  or input('Roster path: ').strip()

    if not args.out:
        stem = Path(roster_path).stem
        out_path = str(Path(roster_path).parent / f'Workday_Change_Reconciliation_{stem}.xlsx')
    else:
        out_path = args.out

    fields_filter = None
    if args.fields:
        fields_filter = {f.strip() for f in args.fields.split(',')}

    print(f'\nLoading roster:        {roster_path}')
    roster = load_roster(roster_path)
    print(f'  → {len(roster):,} associates')

    print(f'Loading change report: {changes_path}')
    win_changes = load_changes(changes_path)
    print(f'  → {len(win_changes)} unique WINs')

    # Resolve snapshot date
    if args.snapshot_date:
        try:
            snapshot_date = datetime.strptime(args.snapshot_date, '%Y-%m-%d').date()
        except ValueError:
            print(f'⚠️  Could not parse --snapshot-date "{args.snapshot_date}". Ignoring.')
            snapshot_date = None
    else:
        snapshot_date = resolve_snapshot_date(roster, roster_path)

    if snapshot_date:
        print(f'  → Snapshot date: {snapshot_date}')
    else:
        print(f'  → Snapshot date: not resolved (⏩ Pending classification disabled)')

    print('\nRunning field-level reconciliation...')
    results = reconcile(
        roster, win_changes,
        snapshot_date=snapshot_date,
        include_location=not args.no_location,
        fields_filter=fields_filter,
    )

    matched = sum(1 for r in results if r['Status'] == '✅ FULLY APPLIED')
    partial  = sum(1 for r in results if '⚠️' in r['Status'])
    pending  = sum(1 for r in results if '🕐' in r['Status'])
    missing  = sum(1 for r in results if '❌' in r['Status'])
    genuine  = sum(
        1 for r in results
        if any(f['match'] == '⚠️' and 'Location' not in f['field'] for f in r.get('fields', []))
    )
    propagation_pct = f"{matched / (matched + partial):.0%}" if (matched + partial) > 0 else 'N/A'

    print(f'\n{"="*56}')
    print(f'  RECONCILIATION COMPLETE  |  Snapshot: {snapshot_date or "unknown"}')
    print(f'{"="*56}')
    print(f'  Total WINs        : {len(results)}')
    print(f'  ✅ Fully applied  : {matched}   ({propagation_pct} propagation rate)')
    print(f'  ⚠️  Partial        : {partial}')
    print(f'  🕐 Pending         : {pending}')
    print(f'  ❌ Not in roster   : {missing}')
    print(f'  Genuine mismatches: {genuine}')
    print(f'{"="*56}')

    if genuine > 0:
        print('\nTOP ISSUES:')
        shown = 0
        for r in results:
            for f in r.get('fields', []):
                if f['match'] == '⚠️' and 'Location' not in f['field']:
                    sev = SEVERITY.get(f['field'], '🟡 Medium')
                    print(f"  {sev}  WIN {r['WIN']} {r.get('Name','')} — {f['field'].split(' → ')[0]}")
                    print(f"         Proposed: {f['proposed'][:60]}")
                    print(f"         Roster:   {f['roster'][:60]}")
                    shown += 1
                    if shown >= 10:
                        break
            if shown >= 10:
                break

    print(f'\nGenerating report...')
    build_report(results, out_path, snapshot_date=snapshot_date)
    print(f'✅ Saved → {out_path}')


if __name__ == '__main__':
    main()
